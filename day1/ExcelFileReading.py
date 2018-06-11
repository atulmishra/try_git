# -*- coding: utf-8 -*-
"""
Created on Mon Jun 11 15:02:00 2018

@author: Administrator
"""
from openpyxl import load_workbook


filePath="D:/citiMLJune2018/day1/GDP.xlsx"

wb = load_workbook(filePath,read_only=True,data_only=True)
sheetRef=wb.get_sheet_by_name("GDP")

for row in range(6,70):
     for col in range (1,6):
         print(sheetRef.cell(column=col,row=row).value,end="\t")